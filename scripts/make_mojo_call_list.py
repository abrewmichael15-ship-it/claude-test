#!/usr/bin/env python3
"""
Reshape a skip-traced leads workbook (e.g. the Upwork-sourced xlsx, sheet
'Seller') into a Mojo Dialer-ready call list - one phone number per
column, all in a single consistent digits-only format (Mojo's own
requirement), one row per contact.

Column names aren't a locked Mojo spec (their import screen maps your
headers to its fields, same as FUB/Click2Mail) - verify against the
actual import screen before your first real call session.

IMPORTANT: this script's INPUT contains real people's phone numbers and
emails. Its OUTPUT does too. Never commit either to git - this repo's
constraints are public-record data only. Run this locally and keep the
output file outside the repo (default: alongside the input file), and
do not call/text anyone on this list until the numbers have been
through a DNC/litigator scrub (see README - Follow Up Boss automation
notes / your chat history with Claude on this).

Usage:
    python3 scripts/make_mojo_call_list.py <input.xlsx> [output.csv]
"""
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

PHONE_COLUMNS_BY_HEADER = [
    "Wireless 1", "Wireless 2", "Wireless 3", "Wireless 4",
    "Landline 1", "Landline 2", "Landline 3", "Landline 4", "Landline 5",
]
EMAIL_COLUMNS_BY_HEADER = ["Email 1", "Email 2"]


def digits_only(phone):
    if not phone:
        return None
    d = re.sub(r"\D", "", str(phone))
    return d if len(d) == 10 else (d[-10:] if len(d) == 11 and d.startswith("1") else None)


def main():
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    in_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else in_path.with_name(
        in_path.stem + "_mojo_call_list.csv"
    )

    wb = openpyxl.load_workbook(in_path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.strip().lower() == "seller"), wb.sheetnames[0])
    ws = wb[sheet_name]

    header_row = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(header_row)}

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw[0]:
            continue

        def get(header):
            i = col_idx.get(header)
            return raw[i] if i is not None and i < len(raw) else None

        csz = str(get("Property City, Sate, Zip Code") or "")
        city, _, state_zip = csz.partition(",")
        state_zip = state_zip.strip()
        state = state_zip[:2] if state_zip else ""
        zip_code = state_zip[2:].strip() if state_zip else ""

        phones = [digits_only(get(h)) for h in PHONE_COLUMNS_BY_HEADER]
        phones = [p for p in phones if p]
        emails = [str(get(h)).strip() for h in EMAIL_COLUMNS_BY_HEADER if get(h)]

        owner_name = str(get("Reported Owner ") or "").replace("\n", " ").strip()

        row = {
            "Name": owner_name,
            "Property_Address": str(get("Property Address") or "").strip(),
            "Property_City": city.strip(),
            "Property_State": state,
            "Property_Zip": zip_code,
        }
        for i, p in enumerate(phones, start=1):
            row[f"Phone{i}"] = p
        for i, e in enumerate(emails, start=1):
            row[f"Email{i}"] = e
        rows.append(row)

    max_phones = max((sum(1 for k in r if k.startswith("Phone")) for r in rows), default=0)
    max_emails = max((sum(1 for k in r if k.startswith("Email")) for r in rows), default=0)
    column_order = (
        ["Name", "Property_Address", "Property_City", "Property_State", "Property_Zip"]
        + [f"Phone{i}" for i in range(1, max_phones + 1)]
        + [f"Email{i}" for i in range(1, max_emails + 1)]
    )
    df = pd.DataFrame(rows).reindex(columns=column_order)
    df.to_csv(out_path, index=False)
    print(f"Wrote {len(df)} rows to {out_path}")
    print()
    print("REMINDER: scrub every phone number against DNC + litigator lists")
    print("before calling or texting anyone on this list. Do not commit")
    print("this file (or the source workbook) to git.")


if __name__ == "__main__":
    main()
