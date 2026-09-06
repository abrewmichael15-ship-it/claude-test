#!/usr/bin/env python3
"""
Cross-reference a skip-traced leads workbook (e.g. the Upwork purchase)
against data/processed/prospects_scored.csv by situs address, so you can
see which of your own scored parcels already have a phone number on
file - useful once equity/violations weights go live and today's
low-scoring matches move up.

This does NOT modify prospects_scored.csv or anything in data/processed
- it writes a small standalone file with parcel_id/rank/score plus the
match's first phone/email, to a path OUTSIDE the repo (default:
alongside the input workbook). Never commit this output.

Usage:
    python3 scripts/flag_purchased_contacts.py <input.xlsx> [output.csv]
"""
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent.parent
PROCESSED_DIR = REPO_ROOT / "data" / "processed"
SCORED_CSV = PROCESSED_DIR / "prospects_scored.csv"


def norm(addr):
    return re.sub(r"[^A-Z0-9 ]", "", str(addr).upper()).strip()


def main():
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    in_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else in_path.with_name(
        in_path.stem + "_matched_to_farm.csv"
    )

    if not SCORED_CSV.exists():
        raise SystemExit(f"Missing {SCORED_CSV} - run build_pipeline.py + score_prospects.py first.")

    wb = openpyxl.load_workbook(in_path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.strip().lower() == "seller"), wb.sheetnames[0])
    ws = wb[sheet_name]
    header_row = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(header_row)}

    leads = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw[0]:
            continue

        def get(header):
            i = col_idx.get(header)
            return raw[i] if i is not None and i < len(raw) else None

        leads.append(
            {
                "lead_property_address": str(get("Property Address") or "").strip(),
                "lead_addr_norm": norm(get("Property Address")),
                "lead_phone1": get("Wireless 1") or get("Landline 1"),
                "lead_email1": get("Email 1"),
            }
        )

    df = pd.read_csv(SCORED_CSV)
    df["situs_norm"] = df["situs_address"].fillna("").apply(norm)

    matches = []
    for lead in leads:
        hit = df[df["situs_norm"] == lead["lead_addr_norm"]]
        if len(hit):
            row = hit.iloc[0]
            matches.append(
                {
                    "parcel_id": row["parcel_id"],
                    "situs_address": row["situs_address"],
                    "municipality": row["municipality"],
                    "rank": row["rank"],
                    "propensity_score": row["propensity_score"],
                    "absentee_owner": row["absentee_owner"],
                    "in_current_top_500": row["rank"] <= 500,
                    "purchased_lead_phone": lead["lead_phone1"],
                    "purchased_lead_email": lead["lead_email1"],
                }
            )

    out = pd.DataFrame(matches).sort_values("rank")
    out.to_csv(out_path, index=False)
    print(f"Matched {len(out)} of {len(leads)} purchased leads to your scored dataset.")
    print(f"Wrote {out_path}")
    print()
    print("REMINDER: this file carries phone/email PII - keep it out of git.")


if __name__ == "__main__":
    main()
